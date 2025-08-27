import streamlit as st
import PyPDF2
import pandas as pd
import re
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import get_column_letter
import random

floor_map = {
    'BF': 'Basement', 'GF': 'Ground Floor', 'GM': 'Ground Mezanin',
    'FF': 'First Floor', 'FM': 'First Mezanin', 'SF': 'Second Floor',
    'SM': 'Second Mezanin', 'TF': 'Third Floor', 'TM': 'Third Mezanin',
    'OF': 'Fourth Floor', 'IF': 'Fifth Floor', 'XF': 'Sixth Floor',
    'F07': 'Seventh Floor', 'F08': 'Eighth Floor', 'F09': 'Ninth Floor'
}

floor_symbol_map = {
    'Basement': 'B', 'Bs': 'B', 'Ground Floor': 'GF', 'Ground Mezanin': 'GM',
    'First Floor': '1', 'First Mezanin': '1M', 'Second Floor': '2', 'Second Mezanin': '2M',
    'Third Floor': '3', 'Third Mezanin': '3M', 'Fourth Floor': '4', 'Fifth Floor': '5',
    'Sixth Floor': '6', 'Seventh Floor': '7', 'Eighth Floor': '8', 'Ninth Floor': '9',
    'Rf': 'R', 'Roof Floor': 'R'
}

def extract_data_from_pdf(pdf_file):
    data = []
    pdf_reader = PyPDF2.PdfReader(pdf_file)
    for page in pdf_reader.pages:
        text = page.extract_text()
        if not text:
            continue
        wo_match = re.search(r'WORKORDER\s*#\s*:\s*(\d+)', text)
        wo_num = wo_match.group(1) if wo_match else ""
        jp_match = re.search(r'JP Code\s*:\s*(\S+)', text)
        check_type = jp_match.group(1).strip()[-1] if jp_match else ""
        qty_match = re.search(r'Asset QTY\s*:\s*(\d+)', text)
        qty = qty_match.group(1) if qty_match else ""
        date_match = re.search(r'Scheduel Start\s*:\s*([\w]+\s+\d{1,2},\s+\d{4})', text)
        formatted_date = ""
        if date_match:
            try:
                date_obj = datetime.strptime(date_match.group(1), '%b %d, %Y')
                formatted_date = date_obj.strftime('%d-%b-%Y')
            except:
                formatted_date = date_match.group(1)
        zone_floor_match = re.search(r'Zone#(\d+),\s*([^,]+)\s*Asset QTY', text)
        if zone_floor_match:
            zone = zone_floor_match.group(1)
            floor = zone_floor_match.group(2).strip()
        else:
            loc_match = re.search(r'Location Code\s*:\s*(\S+)', text)
            loc_code = loc_match.group(1) if loc_match else ""
            zone = loc_code[8:10] if len(loc_code) >= 10 else ""
            floor_code = loc_code[10:12].upper() if len(loc_code) > 10 else ""
            corrections = {'F0': 'OF', '0F': 'OF', 'IF': 'IF', '1F': 'IF'}
            floor_code = corrections.get(floor_code, floor_code)
            floor = floor_map.get(floor_code, floor_code)
        if all([wo_num, zone, floor, qty, check_type, formatted_date]):
            data.append({
                'Work Order': wo_num,
                'Zone': zone,
                'Floor': floor,
                'Quantity': qty,
                'Equipment': 'FHC&PIPE',
                'Type of Check': check_type,
                'Date': formatted_date
            })
    return data

def style_excel(filename):
    wb = load_workbook(filename)
    ws = wb.active
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for col in range(1, ws.max_column + 1):
        max_length = 0
        col_letter = get_column_letter(col)
        for cell in ws[col_letter]:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 5
    dates = {}
    fill_colors = [
        "FFC7CE", "C6EFCE", "FFEB9C", "9CC3E6", "F4CCCC", "D9EAD3",
        "FFE699", "D0E0E3", "FCE5CD", "D9D2E9"
    ]
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        date_cell = row[6]
        date_value = date_cell.value
        if date_value not in dates:
            dates[date_value] = random.choice(fill_colors)
        fill_color = dates[date_value]
        for cell in row:
            cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    wb.save(filename)

def run_project2():
    uploaded_files = st.file_uploader("اختر ملفات PDF للشامية والخدمات والساحات والأنفاق", type=["pdf"], accept_multiple_files=True)
    if uploaded_files:
        all_data = []
        for file in uploaded_files:
            all_data.extend(extract_data_from_pdf(file))
        if all_data:
            df = pd.DataFrame(all_data)
            df['Date'] = pd.to_datetime(df['Date'], format='%d-%b-%Y', errors='coerce')
            df = df.sort_values('Date')
            df['Date'] = df['Date'].dt.strftime('%d-%b-%Y')
            df['Floor'] = df['Floor'].str.strip().str.title()
            df['Floor'] = df['Floor'].map(floor_symbol_map).fillna(df['Floor'])
            output_file = "Extracted_Data_All_Styled.xlsx"
            df.to_excel(output_file, index=False)
            style_excel(output_file)
            with open(output_file, "rb") as f:
                st.download_button(
                    label="⬇️ تحميل ملف Excel",
                    data=f,
                    file_name=output_file,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            st.success("✅ تم استخراج البيانات بنجاح!")
        else:
            st.warning("❌ لم يتم العثور على بيانات صالحة في الملفات.")
