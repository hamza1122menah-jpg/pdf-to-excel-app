import streamlit as st
import fitz  # PyMuPDF
import PyPDF2
import re
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import get_column_letter
import random
import io

# -----------------------------
# خريطة الأدوار والرموز للمشروعين
floor_symbol_map = {
    'Basement': 'B',
    'Bs': 'B',
    'Ground Floor': 'GF',
    'Ground Mezanin': 'GM',
    'First Floor': '1',
    'First Mezanin': '1M',
    'Second Floor': '2',
    'Second Mezanin': '2M',
    'Third Floor': '3',
    'Third Mezanin': '3M',
    'Fourth Floor': '4',
    'Fifth Floor': '5',
    'Sixth Floor': '6',
    'Seventh Floor': '7',
    'Eighth Floor': '8',
    'Ninth Floor': '9',
    'Rf': 'R',
    'Roof Floor': 'R'
}

# -----------------------------
# دالة المشروع الأول
def extract_project1(pdf_file):
    doc = fitz.open(stream=pdf_file.read(), filetype="pdf")
    rows = []
    for page in doc:
        text = page.get_text("text")
        mataf_line = ""
        for line in text.split('\n'):
            if "Mataf Building Project" in line and "Phase" in line:
                mataf_line = line
                break
        workorder = re.search(r'WORKORDER\s*#\s*:\s*(\d+)', text)
        floor_match = re.search(r'Mataf Building Project\s*,([^,]+),([^,]+)', mataf_line)
        phase = re.search(r'Phase\s*#\s*(\d+)', mataf_line) if mataf_line else None
        column = re.search(r'Column\s*([A-Z0-9]+)', mataf_line)
        axis = re.search(r'Axis\s*([A-Z0-9]+)', mataf_line)
        qty = re.search(r'Asset QTY\s*:\s*(\d+)', text)
        type_check = re.search(r'JP Code\s*:\s*([A-Z0-9\-]+)', text)
        date_match = re.search(r'Scheduel Start\s*:\s*([A-Za-z]+\s+\d{1,2},\s+\d{4})', text)

        last_letter = ""
        if type_check:
            match = re.search(r'([A-Z])$', type_check.group(1))
            if match:
                last_letter = match.group(1)
        floor_name = floor_match.group(2).strip() if floor_match else ""
        rows.append({
            "workorder num": workorder.group(1) if workorder else "",
            "Floor": floor_name,
            "phase": phase.group(1) if phase else "",
            "Column": column.group(1) if column else "",
            "Axis": axis.group(1) if axis else "",
            "Quantity": qty.group(1) if qty else "",
            "Equipment": "FHC",
            "Type of check": last_letter,
            "Date": date_match.group(1) if date_match else ""
        })
    return rows

# -----------------------------
# دالة المشروع الثاني
def extract_project2(pdf_file):
    data = []
    reader = PyPDF2.PdfReader(pdf_file)
    for page in reader.pages:
        text = page.extract_text()
        if not text:
            continue
        wo_match = re.search(r'WORKORDER\s*#\s*:\s*(\d+)', text)
        wo_num = wo_match.group(1) if wo_match else ""
        jp_match = re.search(r'JP Code\s*:\s*(\S+)', text)
        check_type = jp_match.group(1).strip()[-1] if jp_match else ""
        qty_match = re.search(r'Asset QTY\s*:\s*(\d+)', text)
        qty = qty_match.group(1) if qty_match else ""
        date_match = re.search(r'Scheduel Start\s*:\s*([\w]+\s+\d{1,2},\s+\d{4})', text)
        formatted_date = ""
        if date_match:
            try:
                date_obj = datetime.strptime(date_match.group(1), '%b %d, %Y')
                formatted_date = date_obj.strftime('%d-%b-%Y')
            except:
                formatted_date = date_match.group(1)
        zone_floor_match = re.search(r'Zone#(\d+),\s*([^,]+)\s*Asset QTY', text)
        if zone_floor_match:
            zone = zone_floor_match.group(1)
            floor = zone_floor_match.group(2).strip()
        else:
            floor = ""
            zone = ""
        if all([wo_num, zone, floor, qty, check_type, formatted_date]):
            data.append({
                'Work Order': wo_num,
                'Zone': zone,
                'Floor': floor,
                'Quantity': qty,
                'Equipment': 'FHC&PIPE',
                'Type of Check': check_type,
                'Date': formatted_date
            })
    return data

# -----------------------------
# دالة تنسيق Excel مشتركة
def style_excel(filename):
    wb = load_workbook(filename)
    ws = wb.active
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for col in range(1, ws.max_column + 1):
        max_length = max(len(str(cell.value)) for cell in ws[get_column_letter(col)] if cell.value) + 5
        ws.column_dimensions[get_column_letter(col)].width = max_length
    # تلوين حسب التاريخ
    dates = {}
    fill_colors = [
        "FFC7CE", "C6EFCE", "FFEB9C", "9CC3E6", "F4CCCC", "D9EAD3",
        "FFE699", "D0E0E3", "FCE5CD", "D9D2E9"
    ]
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        date_cell = row[6]
        date_value = date_cell.value
        if date_value not in dates:
            dates[date_value] = random.choice(fill_colors)
        fill_color = dates[date_value]
        for cell in row:
            cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    wb.save(filename)

# -----------------------------
# واجهة Streamlit
st.title("PDF to Excel - Multi Projects")
st.write("اختر المشروع ورفع الملفات:")

project_choice = st.radio("اختر المشروع:", ("مشروع 1", "مشروع 2"))

uploaded_file = None
if project_choice == "مشروع 1":
    uploaded_file = st.file_uploader("اختر ملف PDF للمشروع 1", type=["pdf"])
elif project_choice == "مشروع 2":
    uploaded_file = st.file_uploader("اختر ملف PDF للمشروع 2", type=["pdf"])

if uploaded_file:
    if project_choice == "مشروع 1":
        data = extract_project1(uploaded_file)
        df_columns = ["workorder num", "Floor", "phase", "Column", "Axis", "Quantity", "Equipment", "Type of check", "Date"]
    else:
        data = extract_project2(uploaded_file)
        df_columns = ['Work Order', 'Zone', 'Floor', 'Quantity', 'Equipment', 'Type of Check', 'Date']

    if data:
        df = pd.DataFrame(data, columns=df_columns)
        df['Date'] = pd.to_datetime(df['Date'], errors='coerce').dt.date
        df = df.sort_values('Date').reset_index(drop=True)
        df['Floor'] = df['Floor'].str.strip().str.title()
        df['Floor'] = df['Floor'].map(floor_symbol_map).fillna(df['Floor'])

        excel_filename = "Extracted_Workorders.xlsx"
        df.to_excel(excel_filename, index=False)
        style_excel(excel_filename)

        with open(excel_filename, "rb") as f:
            st.download_button("⬇️ تحميل ملف Excel", f, file_name=excel_filename,
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        st.success("✅ تم استخراج البيانات بنجاح!")
    else:
        st.warning("❌ لم يتم العثور على بيانات صالحة في الملف.")
